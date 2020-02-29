import openpyxl, os, shutil
from openpyxl.styles import Font, Border, Side

### openpyxl manual ###
#sheet.cell(row=1, column=2) for serialisation
# sheet.max_row / sheet.max_column
# create_sheet() / sheet.title = 'Spam Spam Spam'
# sheet.row_dimensions[1].height = 70
# sheet.column_dimensions['B'].width = 20

# wb.active = index of sheet to make it the active one / by default index 0 is the active sheet


##################################  VARIABLES  ##################################################
x = 1

##################################  FUNCTIONS  ##################################################

def createFolder(path):
    ''' creates folder if it doesn't already exist '''
    try:
        if not os.path.exists(path):
            os.mkdir(path)
    except OSError:
        print('Error creating directory' + path)

############################ EXCEL ############################
createFolder('.\\Folder')

try:
    wb = openpyxl.load_workbook('.\\Folder\\Excel.xlsx')
except FileNotFoundError:
    wb = openpyxl.Workbook()
    ws = wb.active

################ sheet mesi diadromi ################
    ws.title = 'Sheet 1'
    ws['B1'] = 'Example of Table'
    ws.merge_cells('B1:M1') # value and format is defined in its top-left cell
    ws['A2'] = 'DAYS' # or cell.value = whatever
    ws['A3'] = 'Monday'
    ws['A4'] = 'Tuesday'
    ws['A5'] = 'Wednesday'
    ws['A6'] = 'Thursday'
    ws['A7'] = 'Friday'
    ws['A8'] = 'Saturday'
    ws['A9'] = 'Sunday'

    
    medium_border = Border(left=Side(style='medium'), 
                     right=Side(style='medium'), 
                     top=Side(style='medium'), 
                     bottom=Side(style='medium'))
    medium_right_border = Border(right=Side(style='medium'))
    for columnNum in range(1, ws.max_column + 1):
        ws.cell(row=1, column=columnNum).border = medium_border
        ws.cell(row=2, column=columnNum).border = medium_border
    for rowNum in range(1, 10):
        ws.cell(rowNum, column=1).border = medium_right_border
    
    ws.column_dimensions['A'].width = 10.86
    ws['B2'] =  '1'
    # ws.column_dimensions['Β'].width = 8.43 CRUSHES BADLY
    ws['C2'] =  '2'
    ws['D2'] =  '3'
    ws['E2'] =  '4'
    ws['F2'] =  '5'
    ws['G2'] =  '6'
    ws['H2'] =  '7'
    ws['I2'] =  '8'
    ws['J2'] =  '9'
    ws['K2'] =  '10'
    ws['L2'] =  '11'
    ws['M2'] =  '12'

### KM ###
    ws['B3'] =  '80'
    ws['B4'] =  '90'
    ws['B5'] =  '90'     # tractor 1 Monday to Friday
    ws['B6'] =  '90'
    ws['B7'] =  '85'

    ws['C3'] =  '90'
    ws['C4'] =  '85'
    ws['C5'] =  '90'     # tractor 2 Monday to Friday
    ws['C6'] =  '90'
    ws['C7'] =  '90'

    ws['D3'] =  '160'
    ws['D4'] =  '160'
    ws['D5'] =  '160'     # tractor 3 Monday to Friday
    ws['D6'] =  '160'
    ws['D7'] =  '150'

    ws['E3'] =  '65'
    ws['E4'] =  '60'
    ws['E5'] =  '70'     # tractor 4 Monday to Friday
    ws['E6'] =  '65'
    ws['E7'] =  '65'

    ws['F3'] =  '100'
    ws['F4'] =  '100'
    ws['F5'] =  '110'     # tractor 5 Monday to Friday
    ws['F6'] =  '75'
    ws['F7'] =  '70'

    ws['G3'] =  '200'
    ws['G4'] =  '160'
    ws['G5'] =  '180'     # tractor 6 Monday to Friday
    ws['G6'] =  '180'
    ws['G7'] =  '210'

    ws['H3'] =  '120'
    ws['H4'] =  '120'
    ws['H5'] =  '130'     # tractor 7 Monday to Friday
    ws['H6'] =  '120'
    ws['H7'] =  '100'

    ws['I3'] =  '60'
    ws['I4'] =  '100'
    ws['I5'] =  '70'     # tractor 8 Monday to Friday
    ws['I6'] =  '70'
    ws['I7'] =  '70'

    ws['J3'] =  '90'
    ws['J4'] =  '100'
    ws['J5'] =  '100'     # tractor 9 Monday to Friday
    ws['J6'] =  '100'
    ws['J7'] =  '100'

    ws['K3'] =  '120'
    ws['K4'] =  '130'
    ws['K5'] =  '130'     # tractor 10 Monday to Friday
    ws['K6'] =  '120'
    ws['K7'] =  '130'

    ws['L3'] =  '65'
    ws['L4'] =  '65'
    ws['L5'] =  '65'     # tractor 11 Monday to Friday
    ws['L6'] =  '60'
    ws['L7'] =  '60'

    ws['M3'] =  '140'
    ws['M4'] =  '160'
    ws['M5'] =  '160'     # tractor 12 Monday to Friday
    ws['M6'] =  '150'
    ws['M7'] =  '150'
### end KM ###

# ################## styles ##################
#     Arial_11_Font = Font(name='Arial', size=11)

#     Arial_11_bold_Font = Font(name='Arial', size=11, bold=True)
#     for columnNum in range(1, ws.max_column + 1):
#         ws.cell(row=1, column=columnNum).font = Arial_11_bold_Font
############################################

    # wb.save('.\\Folder\\Excel.xlsx')
############# end sheet 1 ###############



################ sheet weekly ################
    ws = wb.create_sheet(0) # sheet to store weekly tractor dict
    ws.title = 'Weekly'

    ws['A1'] = 'Vehicle' # or cell.value = whatever
    ws.column_dimensions['A'].width = 22.43
    ws['B1'] =  'Kilometers'
    ws.column_dimensions['B'].width = 11
    ws['C1'] = 'EURO'
    ws['D1'] = 'Liters'
    ws['E1'] = 'Lt/100Km'
    ws.column_dimensions['E'].width = 13

################## styles ##################
    ws.freeze_panes = 'A2'
    Arial_11_Font = Font(name='Arial', size=11)

    Arial_11_bold_Font = Font(name='Arial', size=11, bold=True)
    for columnNum in range(1, ws.max_column + 1):
        ws.cell(row=1, column=columnNum).font = Arial_11_bold_Font
############################################

############# end sheet weekly ###############


    ws = wb.create_sheet(0)

    ws.title = 'Current Month'
    ws['A1'] = ' Vehicle' # or cell.value = whatever
    ws.column_dimensions['A'].width = 22.43
    ws['B1'] =  'Final Km'
    ws.column_dimensions['B'].width = 10.71
    ws['C1'] = 'Starting Km'
    ws.column_dimensions['C'].width = 11
    ws['D1'] = 'Km'
    ws.column_dimensions['D'].width = 11
    ws['E1'] = 'Invoice for Gas'
    ws.column_dimensions['E'].width = 22.57
    ws['F1'] = 'EURO'
    ws['G1'] = 'Liter'

################## styles ##################
    ws.freeze_panes = 'A2' # freezes first row
    Arial_11_Font = Font(name='Arial', size=11)

    Arial_11_bold_Font = Font(name='Arial', size=11, bold=True)
    for columnNum in range(1, ws.max_column + 1):
        ws.cell(row=1, column=columnNum).font = Arial_11_bold_Font
############################################
    
    wb.save('.\\Folder\\Excel.xlsx')

###############################################################

sheets = wb.get_sheet_names() # in other PC wb.sheetnames
wb.active = len(sheets) - 1 # active sheet always the last
ws = wb.active


for cell in ws['A']:
    print(cell.value)


# create new sheet
if x == 1:
    ws = wb.create_sheet(0) # The active sheet is the sheet that’s on top when the workbook is opened in Excel.
    ws['A1'] = ' Vehicle' # or cell.value = whatever
    ws.column_dimensions['A'].width = 22.43
    ws['B1'] =  'Final Km'
    ws.column_dimensions['B'].width = 10.71
    ws['C1'] = 'Starting Km'
    ws.column_dimensions['C'].width = 11
    ws['D1'] = 'Km'
    ws.column_dimensions['D'].width = 11
    ws['E1'] = 'Invoice for Gas'
    ws.column_dimensions['E'].width = 22.57
    ws['F1'] = 'EURO'
    ws['G1'] = 'Liter'

################## styles ##################
    ws.freeze_panes = 'A2'
    Arial_11_Font = Font(name='Arial', size=11)

    Arial_11_bold_Font = Font(name='Arial', size=11, bold=True)
    for columnNum in range(1, ws.max_column + 1):
        ws.cell(row=1, column=columnNum).font = Arial_11_bold_Font
############################################
    wb.save('.\\Folder\\Excel.xlsx')


# write something in previous sheet and then go back to last one
if 'x' in ws.cell(row=ws.max_row, column=1).value:
    wb.active = 1
    ws = wb.active
    ws.cell(row=ws.max_row +1, column=1).value = '-'
    wb.save('.\\Folder\\Excel.xlsx')
    sheets = wb.get_sheet_names() # in other PC wb.sheetnames
    wb.active = len(sheets) - 1 # active sheet always the last
    ws = wb.active
    wb.save('.\\Folder\\Excel.xlsx')               
